#!/usr/bin/env python
# -*- coding: utf-8 -*-

import requests
import openpyxl
from datetime import datetime, timedelta, timezone
import sys
from getpass import getpass
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import json
import os
import signal

# 全局中断标志
interrupt_flag = threading.Event()

def signal_handler(signum, frame):
    """处理Ctrl+C中断信号"""
    print("\n\n检测到中断信号(Ctrl+C)...", flush=True)
    interrupt_flag.set()


def setup_proxy(username, password, http_proxy, https_proxy):
    """设置代理配置"""
    if username and password:
        # 在代理URL中添加认证信息
        http_proxy_with_auth = http_proxy.replace('http://', f'http://{username}:{password}@')
        https_proxy_with_auth = https_proxy.replace('http://', f'http://{username}:{password}@')
        proxies = {
            'http': http_proxy_with_auth,
            'https': https_proxy_with_auth
        }
    else:
        proxies = {
            'http': http_proxy,
            'https': https_proxy
        }
    return proxies


def read_npm_packages(excel_file):
    """从Excel文件读取npm库名称列表"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        packages = []
        
        # 读取第一列的所有非空值（跳过表头）
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 如果第一列有值
                packages.append(str(row[0]).strip())
        
        wb.close()
        return packages
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        sys.exit(1)


def get_package_versions(package_name, access_token, proxies):
    """获取npm包的版本信息"""
    url = f"https://registry.npmjs.org/{package_name}"
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Accept': 'application/json'
    }
    
    try:
        response = requests.get(url, headers=headers, proxies=proxies, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"获取包 {package_name} 信息失败: {e}")
        return None


def filter_versions_last_year(package_data):
    """筛选2025年的版本信息"""
    if not package_data or 'time' not in package_data:
        return []
    
    # 2025年1月1日 00:00:00 UTC
    year_2025_start = datetime(2025, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    # 2026年1月1日 00:00:00 UTC
    year_2025_end = datetime(2026, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    versions_info = []
    
    time_info = package_data.get('time', {})
    versions = package_data.get('versions', {})
    
    for version, publish_time in time_info.items():
        if version in ['created', 'modified']:
            continue
        
        try:
            publish_date = datetime.fromisoformat(publish_time.replace('Z', '+00:00'))
            if year_2025_start <= publish_date < year_2025_end:
                version_data = versions.get(version, {})
                versions_info.append({
                    'version': version,
                    'publish_time': publish_time,
                    'description': version_data.get('description', ''),
                    'author': version_data.get('author', {}).get('name', '') if isinstance(version_data.get('author'), dict) else str(version_data.get('author', '')),
                    'dependencies': len(version_data.get('dependencies', {}))
                })
        except Exception as e:
            # 静默处理错误，不打印
            continue
    
    # 按发布时间排序
    versions_info.sort(key=lambda x: x['publish_time'], reverse=True)
    return versions_info


def scan_single_package(package_name, access_token, proxies, lock, progress):
    """扫描单个npm包的版本信息（用于多线程）"""
    # 检查是否收到中断信号
    if interrupt_flag.is_set():
        return package_name, None
    
    try:
        package_data = get_package_versions(package_name, access_token, proxies)
        
        if package_data:
            versions = filter_versions_last_year(package_data)
            result = versions
            status_msg = f"✓ 找到 {len(versions)} 个2025年的版本"
        else:
            result = None
            status_msg = "✗ 获取失败"
        
        # 线程安全地更新进度
        with lock:
            progress['completed'] += 1
            print(f"[{progress['completed']}/{progress['total']}] {package_name}: {status_msg}", flush=True)
        
        return package_name, result
    except Exception as e:
        with lock:
            progress['completed'] += 1
            print(f"[{progress['completed']}/{progress['total']}] {package_name}: ✗ 异常: {e}", flush=True)
        return package_name, None


def save_progress(progress_file, results, scanned_packages, all_packages):
    """保存扫描进度"""
    progress_data = {
        'scanned_packages': scanned_packages,
        'all_packages': all_packages,
        'results': results
    }
    with open(progress_file, 'w', encoding='utf-8') as f:
        json.dump(progress_data, f, ensure_ascii=False, indent=2)


def load_progress(progress_file):
    """加载扫描进度"""
    if not os.path.exists(progress_file):
        return None
    
    try:
        with open(progress_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"加载进度文件失败: {e}")
        return None


def remove_progress_file(progress_file):
    """删除进度文件"""
    if os.path.exists(progress_file):
        try:
            os.remove(progress_file)
        except Exception as e:
            print(f"删除进度文件失败: {e}")


def write_results_to_excel(results, output_file):
    """将扫描结果写入Excel文件"""
    wb = openpyxl.Workbook()
    
    # 第一个sheet：详细版本信息
    ws1 = wb.active
    ws1.title = "详细版本信息"
    
    # 写入表头
    headers = ['包名', '版本', '发布时间', '描述', '作者', '依赖数量']
    ws1.append(headers)
    
    # 写入数据
    for package_name, versions in results.items():
        if versions is None:
            ws1.append([package_name, '查找失败', '', '', '', ''])
        elif not versions:
            ws1.append([package_name, '未找到2025年的版本', '', '', '', ''])
        else:
            for version_info in versions:
                ws1.append([
                    package_name,
                    version_info['version'],
                    version_info['publish_time'],
                    version_info['description'],
                    version_info['author'],
                    version_info['dependencies']
                ])
    
    # 调整列宽
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # 第二个sheet：统计信息
    ws2 = wb.create_sheet(title="版本统计")
    ws2.append(['库名', '2025年发布版本数量'])
    
    # 写入统计数据
    for package_name, versions in results.items():
        if versions is None:
            ws2.append([package_name, '查找失败'])
        else:
            ws2.append([package_name, len(versions)])
    
    # 调整统计页列宽
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    print(f"\n扫描结果已保存到: {output_file}")


def main():
    print("=" * 60)
    print("NPM库版本扫描工具")
    print("=" * 60)
    
    # 1. 获取代理配置
    print("\n请输入代理配置（如果不需要代理，直接按回车跳过）:")
    proxy_username = input("代理用户名: ").strip()
    proxy_password = ""
    if proxy_username:
        proxy_password = getpass("代理密码: ")
    
    http_proxy = input("HTTP Proxy (例如: http://proxy.example.com:8080): ").strip()
    https_proxy = input("HTTPS Proxy (例如: http://proxy.example.com:8080): ").strip()
    
    proxies = None
    if http_proxy or https_proxy:
        proxies = setup_proxy(proxy_username, proxy_password, http_proxy, https_proxy)
        print("✓ 代理配置完成")
    
    # 2. 获取输入文件路径
    print("\n请输入包含npm库名称的Excel文件路径:")
    input_file = input("Excel文件路径: ").strip()
    if not input_file:
        print("错误: 必须提供Excel文件路径")
        sys.exit(1)
    
    # 3. 获取npm access token
    print("\n请输入npm access token:")
    access_token = getpass("Access Token: ").strip()
    if not access_token:
        print("错误: 必须提供npm access token")
        sys.exit(1)
    
    # 4. 读取npm包列表
    print(f"\n正在读取Excel文件: {input_file}")
    packages = read_npm_packages(input_file)
    print(f"✓ 共读取到 {len(packages)} 个npm包")
    
    # 检查是否有未完成的进度
    progress_file = input_file.replace('.xlsx', '-progress.json')
    results = {}
    scanned_packages = []
    
    saved_progress = load_progress(progress_file)
    if saved_progress:
        print(f"\n发现未完成的扫描进度:")
        print(f"  已扫描: {len(saved_progress['scanned_packages'])} 个")
        print(f"  待扫描: {len(packages) - len(saved_progress['scanned_packages'])} 个")
        
        resume = input("\n是否继续上次的扫描? (y/n): ").strip().lower()
        if resume == 'y':
            results = saved_progress['results']
            scanned_packages = saved_progress['scanned_packages']
            packages = [pkg for pkg in packages if pkg not in scanned_packages]
            print(f"✓ 已恢复进度，将继续扫描剩余 {len(packages)} 个npm包")
        else:
            print("✓ 将开始全新扫描")
            remove_progress_file(progress_file)
    
    if not packages:
        print("\n所有包已扫描完成！")
        remove_progress_file(progress_file)
        
        # 输出结果
        output_file = input_file.replace('.xlsx', '-扫描结果.xlsx')
        if output_file == input_file:
            output_file = input_file.replace('.xlsx', '') + '-扫描结果.xlsx'
        write_results_to_excel(results, output_file)
        
        # 统计信息
        total_versions = sum(len(versions) for versions in results.values() if versions is not None)
        failed_count = sum(1 for versions in results.values() if versions is None)
        print("\n" + "=" * 60)
        print("扫描完成!")
        print(f"共扫描 {len(results)} 个npm包")
        print(f"查找失败 {failed_count} 个npm包")
        print(f"找到 {total_versions} 个2025年发布的版本")
        print("=" * 60)
        return
    
    # 5. 扫描每个包的版本信息
    print("\n开始扫描npm包版本信息...")
    print("使用多线程并发扫描，请稍候...")
    print("提示: 按 Ctrl+C 可中断扫描并保存当前进度\n")
    
    all_packages_original = read_npm_packages(input_file)  # 保存完整列表用于进度保存
    lock = threading.Lock()
    progress = {'completed': 0, 'total': len(packages)}
    
    # 设置信号处理器
    signal.signal(signal.SIGINT, signal_handler)
    
    executor = None
    try:
        # 使用线程池并发扫描，max_workers控制并发数
        executor = ThreadPoolExecutor(max_workers=15)
        
        # 提交所有任务
        future_to_package = {
            executor.submit(scan_single_package, package, access_token, proxies, lock, progress): package 
            for package in packages
        }
        
        # 收集结果
        for future in as_completed(future_to_package):
            # 检查中断标志
            if interrupt_flag.is_set():
                print("\n正在停止扫描...", flush=True)
                break
            
            try:
                package_name, result = future.result(timeout=1)
                results[package_name] = result
                scanned_packages.append(package_name)
                
                # 每10个包保存一次进度
                if len(scanned_packages) % 10 == 0:
                    save_progress(progress_file, results, scanned_packages, all_packages_original)
            except Exception as e:
                package_name = future_to_package[future]
                results[package_name] = None
                scanned_packages.append(package_name)
                if not interrupt_flag.is_set():
                    print(f"处理 {package_name} 时发生异常: {e}", flush=True)
        
        # 如果是中断，保存进度
        if interrupt_flag.is_set():
            print("\n正在保存当前进度...", flush=True)
            save_progress(progress_file, results, scanned_packages, all_packages_original)
            print(f"✓ 进度已保存到: {progress_file}")
            print(f"✓ 已扫描 {len(scanned_packages)} 个包")
            print("\n下次运行程序时可以选择继续扫描")
        else:
            # 扫描完成，删除进度文件
            remove_progress_file(progress_file)
        
    finally:
        # 确保线程池被正确关闭
        if executor:
            print("\n正在关闭线程池...", flush=True)
            executor.shutdown(wait=False, cancel_futures=True)
    
    # 如果是中断，退出
    if interrupt_flag.is_set():
        sys.exit(0)
    
    # 6. 输出结果到Excel
    print("\n正在生成结果文件...")
    output_file = input_file.replace('.xlsx', '-扫描结果.xlsx')
    if output_file == input_file:
        output_file = input_file.replace('.xlsx', '') + '-扫描结果.xlsx'
    
    write_results_to_excel(results, output_file)
    
    # 统计信息
    total_versions = sum(len(versions) for versions in results.values() if versions is not None)
    failed_count = sum(1 for versions in results.values() if versions is None)
    print("\n" + "=" * 60)
    print("扫描完成!")
    print(f"共扫描 {len(packages)} 个npm包")
    print(f"查找失败 {failed_count} 个npm包")
    print(f"找到 {total_versions} 个2025年发布的版本")
    print("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n程序执行出错: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
